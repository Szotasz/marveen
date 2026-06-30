#!/usr/bin/env python3
"""Rebuild KEF telephely import: per-partner files, clean ASCII names, Attila's
2026-06-28 merge decisions applied. Source of truth = aggregate premises xlsx
(Telephelykapcsolat:1 carries the partner key)."""
import os, re, unicodedata
from collections import defaultdict, OrderedDict
from openpyxl import load_workbook, Workbook

AGG = 'store/kef-telephely/kef-telephely-fa_premises.xlsx'
PARTNERS = 'store/kef-telephely/partner-import/kef-partnerek_partner-import.xlsx'
OUT = 'store/kef-telephely/telephelyek-clean'
HEAD = ['Név','Irányítószám','Település','Utca','Utca típus','Házszám',
        'Telephelykapcsolat: 1','Telephelykapcsolat: 2','Telephelykapcsolat: 3',
        'Telephelykapcsolat: 4','Telephelykapcsolat: 5']
HUN = {'á':'a','é':'e','í':'i','ó':'o','ö':'o','ő':'o','ú':'u','ü':'u','ű':'u',
       'Á':'A','É':'E','Í':'I','Ó':'O','Ö':'O','Ő':'O','Ú':'U','Ü':'U','Ű':'U'}

def tl(s):
    s = ''.join(HUN.get(c, c) for c in (s or ''))
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()

def norm(s):
    s = re.sub(r'\s+', ' ', tl(s).lower()).strip()
    return re.sub(r'\s*-\s*', '-', s)   # "Bács- Kiskun" == "Bács-Kiskun"

def fname(s):
    s = tl(s)
    s = re.sub(r'[^A-Za-z0-9]+', '_', s).strip('_')
    return (s[:90] or 'partner') + '.xlsx'

# canonical 58
wb = load_workbook(PARTNERS, data_only=True); ws = wb.active
canon = [str(r[0]).strip() for r in list(ws.iter_rows(values_only=True))[1:] if r[0] and str(r[0]).strip()]
cnorm = {norm(c): c for c in canon}

def find(sub):
    s = norm(sub)
    for c in canon:
        if s in norm(c): return c
    return None

def remap(partner, telepules, nev):
    pn, tn, nn = norm(partner), norm(telepules), norm(nev)
    if 'jaszberenyi szent erzsebet' in nn or (('hetenyi' in pn or 'jasz-nagykun' in pn or 'jasz nagykun' in pn) and tn in ('jaszbereny', 'nagykata')):
        return find('Jászberényi Szent Erzsébet')           # Q2
    if 'albert schweitzer' in pn or 'albert schweitzer' in nn:
        return find('Albert Schweitzer')                    # auto-merge (Hatvan)
    if 'keszthely' in pn:
        return find('Zala Vármegyei Szent Rafael')          # Q3
    if 'gyor-moson-sopron' in pn or 'petz aladar' in pn or 'kapuvar' in pn or 'lumniczer' in pn:
        return find('Petz Aladár')                          # Q1 (Kapuvár -> Petz)
    if 'jasz-nagykun' in pn or 'jasz nagykun' in pn or 'hetenyi' in pn:
        return find('Jász-Nagykun-Szolnok Vármegyei Hetényi')
    if pn in cnorm: return cnorm[pn]
    for cn, cname in cnorm.items():
        if pn.startswith(cn) or cn.startswith(pn): return cname
    return None

# read aggregate
wb2 = load_workbook(AGG, data_only=True); ws2 = wb2.active
rows = list(ws2.iter_rows(values_only=True))[1:]
groups = defaultdict(list)
unmatched = []
missing_addr = []
for r in rows:
    addr = [('' if r[i] is None else str(r[i]).strip()) for i in range(0, 6)]
    partner = (str(r[6]).strip() if len(r) > 6 and r[6] else '')
    has_addr = any(addr[1:])
    # drop only the genuine stray (no name, no address, no partner)
    if not has_addr and not partner and addr[0] in ('', '(székhely)'):
        continue
    canonp = remap(partner, addr[2], addr[0])
    if not canonp:
        unmatched.append((partner, addr)); continue
    # build descriptive, unique site name
    nev = addr[0]
    if (not nev) or nev == '(székhely)' or norm(nev) == norm(partner) or norm(nev) == norm(canonp):
        site = ' '.join(x for x in [addr[2], addr[3], addr[4], addr[5]] if x).strip()
        nev = f"{canonp} - {site}".strip(' -')
    groups[canonp].append([nev, addr[1], addr[2], addr[3], addr[4], addr[5]])
    if not has_addr:
        missing_addr.append((canonp, nev))

# write files
if os.path.isdir(OUT):
    for f in os.listdir(OUT): os.remove(os.path.join(OUT, f))
os.makedirs(OUT, exist_ok=True)
total = 0
for partner in sorted(groups):
    wb3 = Workbook(); wsx = wb3.active
    wsx.append(HEAD)
    for row in groups[partner]:
        wsx.append(row + ['', '', '', '', ''])
        total += 1
    wb3.save(os.path.join(OUT, fname(partner)))

print(f"canonical partners total: {len(canon)}")
print(f"partners WITH telephely (files written): {len(groups)}")
print(f"telephely rows total: {total}")
zero = [c for c in canon if c not in groups]
print(f"partners with ZERO telephely ({len(zero)}): {zero}")
print(f"UNMATCHED rows: {len(unmatched)}")
for p, a in unmatched: print("   UNMATCHED:", p, '|', a)
print(f"\nMISSING-ADDRESS telephelyek ({len(missing_addr)}) -- included with empty address, need fill:")
for c, n in missing_addr: print(f"   [{c}] {n}")
# spot the decision-relevant partners
for key in ['Petz Aladár', 'Jászberényi Szent Erzsébet', 'Zala Vármegyei Szent Rafael', 'Albert Schweitzer', 'Hetényi']:
    c = find(key)
    if c: print(f"  {key} -> {c}: {len(groups.get(c, []))} telephely")
